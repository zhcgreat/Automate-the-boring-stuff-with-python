#! python
# TextFilestoSpreadsheet.py - read in the contents of several text files and insert those contents into a spreadsheet, with one line of text per row.

import openpyxl, os
wb = openpyxl.Workbook()
sheet = wb.active
txtfilenames = os.listdir('.')
num = 1
for txtfilename in txtfilenames:
    if txtfilename.endswith('txt'):
        with open(txtfilename, encoding = "utf-8") as tf:
            sheet.cell(1, num).value = txtfilename
            lines = tf.readlines()
            for line in lines:
                sheet.cell(lines.index(line)+2, num).value = line
        num += 1
wb.save('Txttosheet.xlsx')
