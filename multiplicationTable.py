#! python3
# coding=utf-8
# multiplicationTable.py - takes a number N from the command line and creates an N×N multiplication tabl

import openpyxl
from openpyxl.styles import Font

N = int(input('Enter a number to creat an an N*N multiplication tabl'))
wb = openpyxl.Workbook()
sheet = wb.active
boldfont = Font(bold=True)
for i in range(N):
    sheet['A' + str(i+2)] = i + 1
    sheet['A' + str(i+2)].font = boldfont
    sheet.cell(1, i+2).value = i + 1
    sheet.cell(1, i+2).font = boldfont
    for j in range(N):
        sheet.cell(i+2, j+2).value= str((i+1)*(j+1))
wb.save('multiplicationTable.xlsx')
